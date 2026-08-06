import io
import re
from decimal import Decimal, InvalidOperation
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Brand, Category, Product
from app.schemas import BulkPriceUpdate
from app.services import prices as price_service

router = APIRouter(prefix="/prices", tags=["prices"])


@router.put("/bulk")
async def bulk_price_update(
    data: BulkPriceUpdate,
    db: AsyncSession = Depends(get_db),
):
    """
    Bulk update selling prices by percentage.
    Must specify either brand_id or category_id (not both).
    Only selling_price is updated, never cost_price.
    """
    if data.brand_id is None and data.category_id is None:
        raise HTTPException(
            status_code=400,
            detail="Must specify either brand_id or category_id",
        )
    if data.brand_id is not None and data.category_id is not None:
        raise HTTPException(
            status_code=400,
            detail="Specify only one of brand_id or category_id, not both",
        )

    if data.brand_id is not None:
        count = await price_service.bulk_update_by_brand(db, data.brand_id, data.percentage)
    else:
        count = await price_service.bulk_update_by_category(db, data.category_id, data.percentage)

    return {"updated": count, "percentage": data.percentage}


def _parse_price(text: str) -> Optional[Decimal]:
    """Parse '$ 46.000,00' or '46000' to Decimal."""
    if not text or not text.strip():
        return None
    s = text.strip().replace("$", "").replace(" ", "").replace("ARS", "")
    if not s or "#REF" in s.upper() or "N/A" in s.upper() or "NO HAY" in s.upper():
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s[:s.rfind(",")].replace(".", "") + "." + s[s.rfind(",") + 1:]
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


@router.post("/upload-excel")
async def upload_excel_prices(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Upload an Excel file to update product prices by name/SKU."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo archivos .xlsx o .xls")

    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {e}")

    results = {"actualizados": 0, "no_encontrados": 0, "errores": [], "detalle": []}

    # Try each sheet
    for ws in wb.worksheets:
        if not hasattr(ws, "iter_rows"):
            continue

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(c or "").strip().lower() for c in rows[0]]

        # Detect columns
        col_name = None
        col_sku = None
        col_price = None
        for i, h in enumerate(headers):
            if not h:
                continue
            if any(kw in h for kw in ["nombre", "producto", "name", "descripcion", "detalle"]):
                col_name = i
            elif any(kw in h for kw in ["codigo", "sku", "barcode", "code", "cod"]):
                col_sku = i
            elif any(kw in h for kw in ["precio", "price", "lista", "venta", "selling"]):
                col_price = i

        if col_price is None:
            results["errores"].append(f"Hoja '{ws.title}': no se encontro columna de precio")
            continue

        for row_idx, row in enumerate(rows[1:], start=2):
            vals = [str(v or "").strip() for v in row]

            if col_price >= len(vals):
                continue
            price_str = vals[col_price] if col_price < len(vals) else ""
            price = _parse_price(price_str)
            if price is None or price == 0:
                continue

            # Try to find product by SKU first, then by name
            product = None
            if col_sku is not None and col_sku < len(vals) and vals[col_sku]:
                product = (await db.execute(
                    select(Product).where(Product.sku == vals[col_sku])
                )).scalar_one_or_none()

            if product is None and col_name is not None and col_name < len(vals) and vals[col_name]:
                product = (await db.execute(
                    select(Product).where(Product.name.ilike(f"%{vals[col_name]}%"))
                )).scalar_one_or_none()

            if product:
                old_price = product.selling_price
                product.selling_price = price
                results["actualizados"] += 1
                results["detalle"].append({
                    "producto": product.name,
                    "precio_anterior": str(old_price) if old_price else "0",
                    "precio_nuevo": str(price),
                })
            else:
                results["no_encontrados"] += 1

    await db.flush()
    wb.close()

    return {
        "actualizados": results["actualizados"],
        "no_encontrados": results["no_encontrados"],
        "errores": results["errores"][:5],
        "detalle": results["detalle"][:20],
    }
