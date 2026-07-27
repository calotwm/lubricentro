"""
Import ALL brand sheets from LISTA DE PRECIO JULIO 2026.xlsx into lubricentro DB.

Each brand sheet has a DIFFERENT column layout. This script handles each
known brand format via a per-sheet format definition.

Usage:
    python import_xlsx_all.py "<path-to-xlsx>"
"""

import os
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

# ---------------------------------------------------------------------------
# ORM models (standalone copy, same as backend)
# ---------------------------------------------------------------------------
from datetime import datetime
from typing import List

from sqlalchemy import (Boolean, DateTime, ForeignKey, Integer,
                        Numeric, String, Text, func)
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship


class Base(DeclarativeBase):
    pass


class Category(Base):
    __tablename__ = "categories"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    name: Mapped[str] = mapped_column(String(100), unique=True, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Brand(Base):
    __tablename__ = "brands"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    name: Mapped[str] = mapped_column(String(100), unique=True, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Product(Base):
    __tablename__ = "products"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    sku: Mapped[Optional[str]] = mapped_column(String(50), unique=True, nullable=True)
    name: Mapped[str] = mapped_column(String(200), nullable=False)
    category_id: Mapped[Optional[int]] = mapped_column(Integer, ForeignKey("categories.id"), nullable=True)
    brand_id: Mapped[Optional[int]] = mapped_column(Integer, ForeignKey("brands.id"), nullable=True)
    specification: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    unit: Mapped[str] = mapped_column(String(20), default="unit")
    cost_price: Mapped[Optional[Decimal]] = mapped_column(Numeric(10, 2), nullable=True)
    selling_price: Mapped[Optional[Decimal]] = mapped_column(Numeric(10, 2), nullable=True)
    current_stock: Mapped[int] = mapped_column(Integer, default=0)
    min_stock: Mapped[int] = mapped_column(Integer, default=0)
    is_active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


# ---------------------------------------------------------------------------
# Price parsing
# ---------------------------------------------------------------------------

def parse_price(text: Any) -> Optional[Decimal]:
    """Parse a price value that could be a number, string with $ and commas, or None."""
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return Decimal(str(round(text, 2)))
    s = str(text).strip()
    if not s or s in ('#REF!', '#N/A', '#VALUE!', 'NO HAY', 'N/A', '-'):
        return None
    # Remove $ and whitespace
    s = s.replace("$", "").replace(" ", "").replace("ARS", "")
    # Handle comma as decimal separator (Argentine format: 46.000,00)
    has_comma = "," in s
    has_dot = "." in s
    if has_comma and has_dot:
        # Could be 46.000,00 (thousands dots, decimal comma) or 1234.56,78
        if s.rfind(",") > s.rfind("."):
            # Decimal comma: 46.000,00 → 46000.00
            int_part = s[:s.rfind(",")].replace(".", "")
            dec_part = s[s.rfind(",") + 1:]
            s = int_part + "." + dec_part
        else:
            # Decimal dot: just remove commas
            s = s.replace(",", "")
    elif has_comma and not has_dot:
        # 1234,56 → 1234.56
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


# ---------------------------------------------------------------------------
# Product name + spec extraction helpers
# ---------------------------------------------------------------------------

CATEGORY_KEYWORDS: list[tuple[str, list[str]]] = [
    ("Aceites de Motor", ["5W30", "5W40", "10W40", "10W60", "15W40", "20W50",
                          "25W60", "0W20", "5W20", "0W30", "0W40", "ACEITE"]),
    ("Transmisiones", ["TRANSMISION", "ATF", "DEX/MERC", "DEXRON", "CVT", "DCT"]),
    ("Aceites de Transmision", ["GEAR", "80/90", "75W90", "85W140", "75W140", "75W80"]),
    ("Aceites de Moto", ["MOTO", "2-STROKE", "2 TIEMPOS", "MOTORCYCLE"]),
    ("Lubricantes Industriales", ["GRASA", "LUBRICANTE", "CHAIN", "CADENA", "WD40"]),
    ("Refrigerantes", ["REFRIGERANTE", "ANTICONGELANTE", "COOLANT"]),
    ("Aditivos y Accesorios", ["ADITIVO", "ADDITIVE", "ZEREX", "LAVA PARABRISAS",
                                "SELLADOR", "STOP LEAK"]),
    ("Filtros de Aceite", ["FILTRO", "FILTER"]),
    ("Baterias", ["BATERIA", "BATTERY", "UB", "WILLARD"]),
    ("Hidraulicos", ["HIDRAULICO", "HYDRAULIC"]),
]


def guess_category(name: str, spec: str = "") -> str:
    """Guess product category from name + spec text."""
    text = f"{name} {spec}".upper()
    for cat, keywords in CATEGORY_KEYWORDS:
        if any(kw.upper() in text for kw in keywords):
            return cat
    return "Sin Categoria"


def extract_capacity(text: str) -> str:
    """Return a cleaned capacity/specification string."""
    if not text:
        return ""
    return str(text).strip()


def build_sku(brand: str, name: str, capacity: str) -> Optional[str]:
    """Build a consistent SKU from brand + name + capacity."""
    parts = [brand[:4].upper()]
    # Take first meaningful chars of name
    name_clean = re.sub(r'[^A-Za-z0-9]', '', name)[:20]
    if name_clean:
        parts.append(name_clean)
    if capacity:
        cap_clean = re.sub(r'[^A-Za-z0-9.]', '', capacity)[:10]
        if cap_clean:
            parts.append(cap_clean)
    return "-".join(parts) if len(parts) > 1 else None


# ---------------------------------------------------------------------------
# Brand-specific format definitions
# ---------------------------------------------------------------------------
# Each entry: (data_start_row, col_name, col_spec, col_list_price, col_cost_price, 
#              price_format, capacity_joined, name_prefix)
# price_format: 'ars' ($ARS with comma decimal), 'number' (plain number)

BRAND_FORMATS: dict[str, dict] = {
    "VALVOLINE": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 3,
        "col_list_price": 4,
        "col_cost_price": 6,
        "price_format": "ars",
        "col_skip_if_empty": 1,  # skip rows without product name
        "notes": "Also col 0 has optional code (BLANCO/GRIS/AZUL/SKU)",
    },
    "TOTAL": {
        "data_start": 3,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 5,  # col 5 seems to be list price (196960)
        "col_cost_price": 2,  # col 2 is cost (15122)
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "MOBIL": {
        "data_start": 3,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 4,  # Guess from similar structure
        "col_cost_price": 2,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "CASTROL": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 7,  # col 7 has prices like 133300 (precio tarjeta?)
        "col_cost_price": 5,  # col 5 has the highest price (212200 = lista)
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "ELF": {
        "data_start": 3,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 4,
        "col_cost_price": 2,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "SHELL": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 5,
        "col_cost_price": 3,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "YPF": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 8,  # col 8 seems to be list price (115770)
        "col_cost_price": 3,  # col 3 is first price column (169258)
        "price_format": "number",
        "col_skip_if_empty": 1,
        "col_code": 0,
    },
    "MOTUL": {
        "data_start": 2,
        "col_name": 1,
        "col_spec": None,  # spec embedded in name
        "col_list_price": 7,  # col 7: 103544 (list price?)
        "col_cost_price": 4,  # col 4: 145000 (first price)
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "TUTELA": {
        "data_start": 2,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 4,
        "col_cost_price": 2,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "BARDAHL": {
        "data_start": 2,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 4,
        "col_cost_price": 2,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "LIQUI MOLY": {
        "data_start": 2,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 4,
        "col_cost_price": 2,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "QUIMBAT": {
        "data_start": 3,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 5,
        "col_cost_price": 3,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "GULF": {
        "data_start": 3,
        "col_name": 0,
        "col_spec": 1,
        "col_list_price": 4,
        "col_cost_price": 2,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
     "WEGA": {
        "data_start": 4,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 6,  # col 6 appears to have price
        "col_cost_price": 4,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "bateria": {
        "data_start": 5,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 3,
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "FILTROS ORIGINALES": {
        "data_start": 4,
        "col_name": 2,  # part description / car model
        "col_spec": 1,  # part number or code
        "col_list_price": 5,  # price column (11000, 23500, etc.)
        "col_cost_price": 3,  # cost column (139.56, 497.9, etc.)
        "price_format": "number",
        "col_skip_if_empty": 2,
    },
    "MANN FILTER": {
        "data_start": 7,
        "col_name": 9,  # PRECIOS DE LISTA SIN IVA column
        "col_spec": 0,  # part reference (0.6, etc.)
        "col_list_price": 5,  # PRECIO GANANCIA column
        "col_cost_price": 2,  # Precio unit. sin IVA
        "price_format": "number",
        "col_skip_if_empty": 9,
    },
    "FRAM": {
        "data_start": 3,
        "col_name": 1,  # Product code/name
        "col_spec": 2,  # Description
        "col_list_price": 5,  # Price column
        "col_cost_price": 4,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "MASTERFILT": {
        "data_start": 3,
        "col_name": 2,  # DESCRIPCION DEL ARTICULO
        "col_spec": 0,  # CODIGO
        "col_list_price": 5,  # column 5 has prices (142, 174, etc.)
        "col_cost_price": 3,  # column 3 has $ (CA2690PL, etc.)
        "price_format": "number",
        "col_skip_if_empty": 2,
    },
    "FAP": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 4,
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "TECNECO": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 4,
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "FARO ": {  # Note: has trailing space in Excel
        "data_start": 11,
        "col_name": 0,  # APPLICATION
        "col_spec": 2,  # FARO code
        "col_list_price": 5,  # PRECIO column
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "FARO": {
        "data_start": 6,
        "col_name": 0,  # APPLICATION / car model
        "col_spec": 2,  # FARO code
        "col_list_price": 5,  # S/IVA price
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 0,
    },
    "MARENO": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 4,
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "VARIOS": {
        "data_start": 3,
        "col_name": 1,
        "col_spec": 2,
        "col_list_price": 3,
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "PUMA": {
        "data_start": 8,
        "col_name": 1,  # DETALLE (10W40, etc.)
        "col_spec": 2,  # CAPACIDAD
        "col_list_price": 5,  # PRECIO column 2
        "col_cost_price": 3,  # PRECIO column 1
        "price_format": "number",
        "col_skip_if_empty": 1,
    },
    "DM": {
        "data_start": 7,
        "col_name": 2,  # DESCRIPCION
        "col_spec": 1,  # CODIGO
        "col_list_price": None,  # No price data
        "col_cost_price": None,
        "price_format": "number",
        "col_skip_if_empty": 2,
    },
}


# ---------------------------------------------------------------------------
# Import a single brand sheet
# ---------------------------------------------------------------------------

def import_brand_sheet(session: Session, brand_name: str,
                       ws: openpyxl.worksheet.worksheet.Worksheet,
                       fmt: dict) -> tuple[int, int, list[str]]:
    """Import products from one brand sheet. Returns (new_count, update_count, warnings)."""
    new_count = 0
    update_count = 0
    warnings = []

    # Get or create brand
    brand = session.execute(select(Brand).where(Brand.name == brand_name)).scalar_one_or_none()
    if brand is None:
        brand = Brand(name=brand_name)
        session.add(brand)
        session.flush()
    brand_id = brand.id

    category_cache: dict[str, int] = {}

    col_name = fmt.get("col_name")
    col_spec = fmt.get("col_spec")
    col_list_price = fmt.get("col_list_price")
    col_cost_price = fmt.get("col_cost_price")
    price_fmt = fmt.get("price_format", "number")
    skip_col = fmt.get("col_skip_if_empty")
    data_start = fmt.get("data_start", 1)

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_num < data_start:
            continue

        row_list = list(row)

        # Check if this row has anything useful
        if skip_col is not None:
            val = str(row_list[skip_col]).strip() if skip_col < len(row_list) else ""
            if not val:
                continue
        elif not any(cell is not None and str(cell).strip() for cell in row_list[:8]):
            continue

        # Extract name
        name = ""
        if col_name is not None and col_name < len(row_list):
            name = str(row_list[col_name] or "").strip()

        if not name or name in ("#REF!", "#N/A", "-"):
            continue

        # Extract specification/capacity
        spec_str = ""
        if col_spec is not None and col_spec < len(row_list):
            spec_str = str(row_list[col_spec] or "").strip()

        # Extract prices
        list_price = None
        if col_list_price is not None and col_list_price < len(row_list):
            list_price = parse_price(row_list[col_list_price])

        cost_price = None
        if col_cost_price is not None and col_cost_price < len(row_list):
            cost_price = parse_price(row_list[col_cost_price])

        # Skip if no valid selling price
        if list_price is None or list_price == 0:
            warnings.append(f"  {brand_name} row {row_num}: '{name[:40]}' — no valid price, skipping")
            continue

        # Build product name with spec
        product_name = name
        if spec_str and spec_str not in name:
            product_name = f"{name} {spec_str}"
        # Remove brand prefix if redundant
        brand_upper = brand_name.upper()
        for prefix in [brand_upper, brand_upper + " ", brand_upper + "/"]:
            if product_name.upper().startswith(prefix):
                product_name = product_name[len(prefix):].strip()
                break
        if not product_name:
            product_name = name

        # Build SKU
        sku = build_sku(brand_name, product_name, spec_str)

        # Guess category
        cat_name = guess_category(product_name, spec_str)
        if cat_name not in category_cache:
            existing_cat = session.execute(select(Category).where(Category.name == cat_name)).scalar_one_or_none()
            if existing_cat is None:
                existing_cat = Category(name=cat_name)
                session.add(existing_cat)
                session.flush()
            category_cache[cat_name] = existing_cat.id
        category_id = category_cache[cat_name]

        # Upsert by SKU or name
        existing = None
        if sku:
            existing = session.execute(select(Product).where(Product.sku == sku)).scalar_one_or_none()
        if existing is None:
            existing = session.execute(
                select(Product).where(Product.name == product_name, Product.brand_id == brand_id)
            ).scalar_one_or_none()

        if existing:
            existing.name = product_name
            existing.selling_price = list_price
            if cost_price is not None:
                existing.cost_price = cost_price
            if spec_str:
                existing.specification = spec_str
            existing.category_id = category_id
            existing.is_active = True
            if sku:
                existing.sku = sku
            update_count += 1
        else:
            prod = Product(
                name=product_name,
                sku=sku,
                selling_price=list_price,
                cost_price=cost_price,
                specification=spec_str or None,
                unit="unit",
                current_stock=0,
                min_stock=0,
                brand_id=brand_id,
                category_id=category_id,
                is_active=True,
            )
            session.add(prod)
            new_count += 1

    return new_count, update_count, warnings


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python import_xlsx_all.py <path-to-xlsx>")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)

    # DB connection
    db_url = os.environ.get("DB_URL", "sqlite:///../backend/lubricentro.db")
    engine = create_engine(db_url, echo=False)
    SessionLocal = sessionmaker(bind=engine)

    print(f"Reading: {xlsx_path}")
    print(f"Database: {db_url}")
    print()

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    all_warnings: list[str] = []
    total_new = 0
    total_updates = 0
    total_brands = 0
    skipped_sheets: list[str] = []

    with SessionLocal() as session:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Accept both Worksheet and ReadOnlyWorksheet
            if not hasattr(ws, 'iter_rows'):
                skipped_sheets.append(f"{sheet_name} (chart/other)")
                continue
            if sheet_name.lower().startswith("inicio") or sheet_name.startswith("Grafico") or sheet_name.startswith("Grafico"):
                skipped_sheets.append(f"{sheet_name} (metadata)")
                continue

            # Check if we have a format for this sheet
            fmt = BRAND_FORMATS.get(sheet_name)
            if fmt is None:
                skipped_sheets.append(f"{sheet_name} (no format defined)")
                continue

            if fmt.get("col_list_price") is None:
                skipped_sheets.append(f"{sheet_name} (complex pricing)")
                continue

            brand_name = sheet_name.strip()
            print(f"Processing: {brand_name} ...", end=" ")

            try:
                new_c, upd_c, warns = import_brand_sheet(session, brand_name, ws, fmt)
                total_new += new_c
                total_updates += upd_c
                total_brands += 1
                all_warnings.extend(warns)
                print(f"{new_c} new, {upd_c} updated")
            except Exception as e:
                print(f"ERROR: {e}")
                all_warnings.append(f"{brand_name}: import failed — {e}")

            # Commit per sheet to avoid losing data on errors
            try:
                session.commit()
            except Exception as e:
                session.rollback()
                print(f"  COMMIT ERROR: {e}")
                all_warnings.append(f"{brand_name}: commit failed — {e}")

    wb.close()

    # Print summary
    print("\n" + "=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)

    # Verify with fresh session
    with SessionLocal() as session:
        brands_count = session.execute(select(Brand)).scalars().all()
        cats_count = session.execute(select(Category)).scalars().all()
        prods_count = session.execute(select(Product)).scalars().all()

        print(f"\nBrands: {len(brands_count)}")
        for b in brands_count:
            cnt = session.execute(select(Product).where(Product.brand_id == b.id)).scalar()
            print(f"  {b.name}: {cnt} products" if cnt else f"  {b.name}: 0 products")

        print(f"\nCategories: {len(cats_count)}: {', '.join(c.name for c in cats_count)}")
        print(f"Total products: {len(prods_count)}")
        print(f"New: {total_new}, Updated: {total_updates}")

    print(f"\nSkipped sheets: {', '.join(skipped_sheets) if skipped_sheets else 'none'}")

    if all_warnings:
        print(f"\nWarnings ({len(all_warnings)}):")
        for w in all_warnings[:20]:
            print(f"  {w}")
        if len(all_warnings) > 20:
            print(f"  ... and {len(all_warnings) - 20} more")


if __name__ == "__main__":
    main()
