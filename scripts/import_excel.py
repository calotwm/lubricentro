"""
Excel-to-SQLite import script for lubricentro.

Reads categories, brands, and products from an Excel workbook and inserts
them into the lubricentro database. Missing categories/brands are created
automatically. Rows with missing required fields are skipped with a warning.

Usage:
    python import_excel.py path/to/file.xlsx

Environment:
    DB_URL  SQLAlchemy connection string (default: sqlite:///../backend/lubricentro.db)
"""

import os
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

# ---------------------------------------------------------------------------
# ORM models (mirrored from backend/app/models.py to keep this script standalone)
# ---------------------------------------------------------------------------
from sqlalchemy import (
    Boolean,
    DateTime,
    ForeignKey,
    Integer,
    Numeric,
    String,
    Text,
    func,
)
from sqlalchemy.orm import (
    DeclarativeBase,
    Mapped,
    mapped_column,
    relationship,
)
from typing import List


class Base(DeclarativeBase):
    pass


class Category(Base):
    __tablename__ = "categories"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    name: Mapped[str] = mapped_column(String(100), unique=True, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now(), nullable=False)
    products: Mapped[List["Product"]] = relationship(back_populates="category", lazy="selectin")


class Brand(Base):
    __tablename__ = "brands"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    name: Mapped[str] = mapped_column(String(100), unique=True, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now(), nullable=False)
    products: Mapped[List["Product"]] = relationship(back_populates="brand", lazy="selectin")


class Product(Base):
    __tablename__ = "products"
    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    sku: Mapped[Optional[str]] = mapped_column(String(50), unique=True, nullable=True)
    name: Mapped[str] = mapped_column(String(200), nullable=False)
    category_id: Mapped[Optional[int]] = mapped_column(Integer, ForeignKey("categories.id"), nullable=True)
    brand_id: Mapped[Optional[int]] = mapped_column(Integer, ForeignKey("brands.id"), nullable=True)
    specification: Mapped[Optional[str]] = mapped_column(Text, nullable=True)
    unit: Mapped[str] = mapped_column(String(20), default="unit", nullable=False)
    cost_price: Mapped[Optional[Decimal]] = mapped_column(Numeric(10, 2), nullable=True)
    selling_price: Mapped[Optional[Decimal]] = mapped_column(Numeric(10, 2), nullable=True)
    current_stock: Mapped[int] = mapped_column(Integer, default=0, nullable=False)
    min_stock: Mapped[int] = mapped_column(Integer, default=0, nullable=False)
    is_active: Mapped[bool] = mapped_column(Boolean, default=True, nullable=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now(), nullable=False)
    category: Mapped[Optional[Category]] = relationship(back_populates="products", lazy="selectin")
    brand: Mapped[Optional[Brand]] = relationship(back_populates="products", lazy="selectin")


# ---------------------------------------------------------------------------
# Sheet-name detection (flexible, case-insensitive, accent-insensitive)
# ---------------------------------------------------------------------------

CATEGORY_ALIASES = {"categorias", "categoria", "categories", "category", "cat"}
BRAND_ALIASES = {"marcas", "marca", "brands", "brand"}
PRODUCT_ALIASES = {"productos", "producto", "products", "product", "prod", "inventario", "inventory"}


def normalize_sheet_name(name: str) -> str:
    """Lowercase + strip accents for matching."""
    import unicodedata
    normalized = unicodedata.normalize("NFD", name.lower().strip())
    return "".join(c for c in normalized if unicodedata.category(c) != "Mn")


def find_sheet(wb: openpyxl.Workbook, aliases: set) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    """Return the first worksheet whose normalized name matches any alias."""
    for ws in wb.worksheets:
        if normalize_sheet_name(ws.title) in aliases:
            return ws
    return None


# ---------------------------------------------------------------------------
# Column-name detection (flexible, case-insensitive, accent-insensitive)
# ---------------------------------------------------------------------------

def normalize_col(name: str) -> str:
    """Lowercase, strip accents, replace spaces/hyphens with underscores."""
    import unicodedata
    normalized = unicodedata.normalize("NFD", name.lower().strip())
    stripped = "".join(c for c in normalized if unicodedata.category(c) != "Mn")
    return stripped.replace(" ", "_").replace("-", "_")


COLUMN_MAP = {
    "name":         {"name", "producto", "product_name", "nombre"},
    "sku":          {"sku", "codigo", "code", "codigo_de_barras", "barcode", "cod"},
    "category":     {"categoria", "category", "cat"},
    "brand":        {"marca", "brand"},
    "selling_price":{"precio", "price", "precio_venta", "selling_price", "precio_publico"},
    "cost_price":   {"costo", "cost", "precio_costo", "cost_price"},
    "current_stock":{"stock", "cantidad", "current_stock", "existencia", "stock_actual"},
    "min_stock":    {"stock_minimo", "min_stock", "minimo", "stock_min"},
    "specification":{"especificacion", "specification", "viscosidad", "viscosity", "spec"},
    "unit":         {"unidad", "unit", "unidad_medida", "uom"},
}


def build_header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """
    Map logical field names to column indices (0-based).
    Returns dict like {"name": 0, "sku": 2, ...}
    """
    header_row = [str(cell.value or "").strip() for cell in ws[1]]
    field_map = {}
    for logical_name, aliases in COLUMN_MAP.items():
        for idx, raw_col in enumerate(header_row):
            if normalize_col(raw_col) in aliases:
                field_map[logical_name] = idx
                break
    return field_map


def cell_value(row: list, col_idx: Optional[int]) -> Optional[str]:
    """Safely extract a cell value from a row by column index."""
    if col_idx is None or col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None:
        return None
    return str(val).strip()


def to_decimal(val: Optional[str]) -> Optional[Decimal]:
    """Convert a string to Decimal, returning None on failure."""
    if val is None or val == "":
        return None
    try:
        # Handle comma as decimal separator (common in Spanish locales)
        cleaned = val.replace(",", ".")
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def to_int(val: Optional[str], default: int = 0) -> int:
    """Convert a string to int, returning default on failure."""
    if val is None or val == "":
        return default
    try:
        return int(float(val.replace(",", ".")))
    except (ValueError, InvalidOperation):
        return default


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

def get_or_create_category(session: Session, name: str) -> int:
    """Look up or create a category by name. Returns the category ID."""
    cat = session.execute(select(Category).where(Category.name == name)).scalar_one_or_none()
    if cat is None:
        cat = Category(name=name)
        session.add(cat)
        session.flush()
    return cat.id


def get_or_create_brand(session: Session, name: str) -> int:
    """Look up or create a brand by name. Returns the brand ID."""
    brand = session.execute(select(Brand).where(Brand.name == name)).scalar_one_or_none()
    if brand is None:
        brand = Brand(name=name)
        session.add(brand)
        session.flush()
    return brand.id


def import_categories(ws: openpyxl.worksheet.worksheet.Worksheet, session: Session) -> tuple:
    """Import categories from a worksheet. Returns (count, warnings)."""
    header_map = build_header_map(ws)
    name_idx = header_map.get("name")
    if name_idx is None:
        # Fallback: try first column
        name_idx = 0

    count = 0
    warnings = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_list = list(row)
        name = cell_value(row_list, name_idx)
        if not name:
            warnings.append(f"Categories row {row_num}: missing name, skipping")
            continue
        existing = session.execute(select(Category).where(Category.name == name)).scalar_one_or_none()
        if existing is None:
            session.add(Category(name=name))
            count += 1
    return count, warnings


def import_brands(ws: openpyxl.worksheet.worksheet.Worksheet, session: Session) -> tuple:
    """Import brands from a worksheet. Returns (count, warnings)."""
    header_map = build_header_map(ws)
    name_idx = header_map.get("name")
    if name_idx is None:
        name_idx = 0

    count = 0
    warnings = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_list = list(row)
        name = cell_value(row_list, name_idx)
        if not name:
            warnings.append(f"Brands row {row_num}: missing name, skipping")
            continue
        existing = session.execute(select(Brand).where(Brand.name == name)).scalar_one_or_none()
        if existing is None:
            session.add(Brand(name=name))
            count += 1
    return count, warnings


def import_products(ws: openpyxl.worksheet.worksheet.Worksheet, session: Session) -> tuple:
    """Import products from a worksheet. Returns (count, warnings)."""
    header_map = build_header_map(ws)
    count = 0
    warnings = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_list = list(row)

        # Required: name
        name = cell_value(row_list, header_map.get("name"))
        if not name:
            warnings.append(f"Row {row_num}: missing name, skipping")
            continue

        # Required: selling_price
        price_str = cell_value(row_list, header_map.get("selling_price"))
        price = to_decimal(price_str)
        if price is None:
            warnings.append(f"Row {row_num}: missing or invalid price, skipping")
            continue

        # Optional fields
        sku = cell_value(row_list, header_map.get("sku"))
        spec = cell_value(row_list, header_map.get("specification"))
        unit = cell_value(row_list, header_map.get("unit")) or "unit"
        cost_price = to_decimal(cell_value(row_list, header_map.get("cost_price")))
        current_stock = to_int(cell_value(row_list, header_map.get("current_stock")), default=0)
        min_stock = to_int(cell_value(row_list, header_map.get("min_stock")), default=0)

        # Category lookup (create if missing)
        category_id = None
        cat_name = cell_value(row_list, header_map.get("category"))
        if cat_name:
            category_id = get_or_create_category(session, cat_name)

        # Brand lookup (create if missing)
        brand_id = None
        brand_name = cell_value(row_list, header_map.get("brand"))
        if brand_name:
            brand_id = get_or_create_brand(session, brand_name)

        # Upsert by SKU if present, otherwise by name
        product = None
        if sku:
            product = session.execute(select(Product).where(Product.sku == sku)).scalar_one_or_none()
        if product is None:
            product = session.execute(select(Product).where(Product.name == name)).scalar_one_or_none()

        if product:
            # Update existing
            product.name = name
            product.selling_price = price
            if sku:
                product.sku = sku
            if spec is not None:
                product.specification = spec
            product.unit = unit
            if cost_price is not None:
                product.cost_price = cost_price
            product.current_stock = current_stock
            product.min_stock = min_stock
            if category_id is not None:
                product.category_id = category_id
            if brand_id is not None:
                product.brand_id = brand_id
        else:
            # Insert new
            product = Product(
                name=name,
                sku=sku,
                selling_price=price,
                cost_price=cost_price,
                current_stock=current_stock,
                min_stock=min_stock,
                specification=spec,
                unit=unit,
                category_id=category_id,
                brand_id=brand_id,
                is_active=True,
            )
            session.add(product)
            count += 1

    return count, warnings


def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py <path/to/file.xlsx>")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"Error: file not found: {excel_path}")
        sys.exit(1)

    # Database connection
    db_url = os.environ.get("DB_URL", "sqlite:///../backend/lubricentro.db")
    engine = create_engine(db_url, echo=False)
    SessionLocal = sessionmaker(bind=engine)

    # Load workbook
    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    all_warnings: list = []
    cat_count = 0
    brand_count = 0
    prod_count = 0

    with SessionLocal() as session:
        # --- Categories ---
        cat_ws = find_sheet(wb, CATEGORY_ALIASES)
        if cat_ws:
            cat_count, warns = import_categories(cat_ws, session)
            all_warnings.extend(warns)
            session.commit()
        else:
            all_warnings.append("No categories sheet found, skipping")

        # --- Brands ---
        brand_ws = find_sheet(wb, BRAND_ALIASES)
        if brand_ws:
            brand_count, warns = import_brands(brand_ws, session)
            all_warnings.extend(warns)
            session.commit()
        else:
            all_warnings.append("No brands sheet found, skipping")

        # --- Products ---
        prod_ws = find_sheet(wb, PRODUCT_ALIASES)
        if prod_ws:
            prod_count, warns = import_products(prod_ws, session)
            all_warnings.extend(warns)
            session.commit()
        else:
            all_warnings.append("No products sheet found, skipping")

    wb.close()

    # Print warnings
    for w in all_warnings:
        print(f"WARNING: {w}")

    # Summary
    print(f"\nImported: {cat_count} categories, {brand_count} brands, {prod_count} products. Warnings: {len(all_warnings)}")


if __name__ == "__main__":
    main()
