"""
Tests for the Excel import script (scripts/import_excel.py).

Creates temporary XLSX files, runs the import functions against an
in-memory SQLite database, and verifies the imported data.

Run with:
    cd backend
    python -m pytest ../scripts/test_import_excel.py -v
"""

import os
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

# Add scripts directory to path so we can import import_excel
SCRIPTS_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SCRIPTS_DIR))

from import_excel import (
    Base,
    Brand,
    Category,
    Product,
    build_header_map,
    find_sheet,
    import_brands,
    import_categories,
    import_products,
    normalize_sheet_name,
)
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker


@pytest.fixture
def db_session():
    """Provide a clean in-memory SQLite session for each test."""
    engine = create_engine("sqlite:///:memory:", echo=False)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)
    with SessionLocal() as session:
        yield session
    Base.metadata.drop_all(engine)
    engine.dispose()


def create_test_workbook(
    categories=None, brands=None, products=None,
    cat_sheet="Categorias", brand_sheet="Marcas", prod_sheet="Productos",
):
    """Create a temporary XLSX workbook with test data.

    Each parameter is a list of dicts. Returns the workbook path.
    """
    wb = openpyxl.Workbook()

    # Categories sheet
    ws_cat = wb.active
    ws_cat.title = cat_sheet
    ws_cat.append(["name"])
    if categories:
        for cat in categories:
            ws_cat.append([cat.get("name", "")])

    # Brands sheet
    ws_brand = wb.create_sheet(brand_sheet)
    ws_brand.append(["name"])
    if brands:
        for brand in brands:
            ws_brand.append([brand.get("name", "")])

    # Products sheet
    ws_prod = wb.create_sheet(prod_sheet)
    ws_prod.append([
        "name", "sku", "category", "brand",
        "selling_price", "cost_price", "current_stock",
        "min_stock", "specification", "unit",
    ])
    if products:
        for prod in products:
            ws_prod.append([
                prod.get("name", ""),
                prod.get("sku", ""),
                prod.get("category", ""),
                prod.get("brand", ""),
                prod.get("selling_price", ""),
                prod.get("cost_price", ""),
                prod.get("current_stock", ""),
                prod.get("min_stock", ""),
                prod.get("specification", ""),
                prod.get("unit", ""),
            ])

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    wb.close()
    return tmp.name


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestNormalizeSheetName:
    def test_basic(self):
        assert normalize_sheet_name("Categorias") == "categorias"

    def test_accents(self):
        assert normalize_sheet_name("Categorías") == "categorias"

    def test_spaces(self):
        assert normalize_sheet_name("  Productos  ") == "productos"


class TestFindSheet:
    def test_finds_matching_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Categorias"
        ws = find_sheet(wb, {"categorias", "categories"})
        assert ws is not None
        assert ws.title == "Categorias"

    def test_returns_none_when_no_match(self):
        wb = openpyxl.Workbook()
        wb.active.title = "SomethingElse"
        ws = find_sheet(wb, {"categorias", "categories"})
        assert ws is None


class TestImportCategories:
    def test_import_categories(self, db_session):
        path = create_test_workbook(
            categories=[{"name": "Aceites"}, {"name": "Filtros"}]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"categorias", "categories"})
            count, warnings = import_categories(ws, db_session)
            db_session.commit()

            assert count == 2
            assert len(warnings) == 0

            cats = db_session.execute(select(Category)).scalars().all()
            names = {c.name for c in cats}
            assert "Aceites" in names
            assert "Filtros" in names
        finally:
            os.unlink(path)

    def test_skip_empty_category_name(self, db_session):
        path = create_test_workbook(
            categories=[{"name": "Aceites"}, {"name": ""}]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"categorias", "categories"})
            count, warnings = import_categories(ws, db_session)
            db_session.commit()

            assert count == 1
            assert len(warnings) == 1
            assert "missing name" in warnings[0]
        finally:
            os.unlink(path)


class TestImportBrands:
    def test_import_brands(self, db_session):
        path = create_test_workbook(
            brands=[{"name": "Motul"}, {"name": "Castrol"}]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"marcas", "brands"})
            count, warnings = import_brands(ws, db_session)
            db_session.commit()

            assert count == 2
            brands = db_session.execute(select(Brand)).scalars().all()
            names = {b.name for b in brands}
            assert "Motul" in names
            assert "Castrol" in names
        finally:
            os.unlink(path)


class TestImportProducts:
    def test_import_products(self, db_session):
        path = create_test_workbook(
            products=[
                {
                    "name": "Aceite 20W-50",
                    "sku": "MOT-001",
                    "category": "Aceites",
                    "brand": "Motul",
                    "selling_price": "100.50",
                    "cost_price": "50.00",
                    "current_stock": "20",
                    "min_stock": "5",
                    "specification": "20W-50",
                    "unit": "unit",
                }
            ]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"productos", "products"})
            count, warnings = import_products(ws, db_session)
            db_session.commit()

            assert count == 1
            assert len(warnings) == 0

            product = db_session.execute(select(Product)).scalar_one()
            assert product.name == "Aceite 20W-50"
            assert product.sku == "MOT-001"
            assert product.selling_price is not None
            assert float(product.selling_price) == 100.50
            assert product.current_stock == 20

            # Category and brand should be auto-created
            cats = db_session.execute(select(Category)).scalars().all()
            assert any(c.name == "Aceites" for c in cats)
            brands = db_session.execute(select(Brand)).scalars().all()
            assert any(b.name == "Motul" for b in brands)
        finally:
            os.unlink(path)

    def test_skip_missing_name(self, db_session):
        """Row with missing name is skipped with a warning."""
        path = create_test_workbook(
            products=[
                {"name": "", "selling_price": "100"},  # no name
                {"name": "Valid Product", "selling_price": "50"},
            ]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"productos", "products"})
            count, warnings = import_products(ws, db_session)
            db_session.commit()

            assert count == 1
            assert len(warnings) == 1
            assert "missing name" in warnings[0]
        finally:
            os.unlink(path)

    def test_skip_missing_price(self, db_session):
        """Row with missing price is skipped with a warning."""
        path = create_test_workbook(
            products=[
                {"name": "No Price", "selling_price": ""},  # no price
                {"name": "Valid", "selling_price": "50"},
            ]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"productos", "products"})
            count, warnings = import_products(ws, db_session)
            db_session.commit()

            assert count == 1
            assert len(warnings) == 1
            assert "price" in warnings[0].lower()
        finally:
            os.unlink(path)

    def test_upsert_by_sku(self, db_session):
        """Importing a product with an existing SKU updates it instead of creating a new one."""
        # First import
        path = create_test_workbook(
            products=[
                {"name": "Aceite 20W-50", "sku": "MOT-001", "selling_price": "100"},
            ]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"productos", "products"})
            import_products(ws, db_session)
            db_session.commit()
        finally:
            os.unlink(path)

        # Second import with same SKU but different price
        path = create_test_workbook(
            products=[
                {"name": "Aceite 20W-50", "sku": "MOT-001", "selling_price": "120"},
            ]
        )
        try:
            wb = openpyxl.load_workbook(path)
            ws = find_sheet(wb, {"productos", "products"})
            count, _ = import_products(ws, db_session)
            db_session.commit()

            # count should be 0 (update, not insert)
            assert count == 0

            product = db_session.execute(select(Product)).scalar_one()
            assert float(product.selling_price) == 120.0
        finally:
            os.unlink(path)


class TestMissingSheet:
    def test_missing_brands_sheet(self, db_session):
        """When brands sheet is missing, find_sheet returns None."""
        wb = openpyxl.Workbook()
        wb.active.title = "Categorias"
        wb.active.append(["name"])
        wb.active.append(["Aceites"])

        # No brands sheet — only categories
        brand_ws = find_sheet(wb, {"marcas", "brands"})
        assert brand_ws is None

    def test_missing_products_sheet(self, db_session):
        """When products sheet is missing, find_sheet returns None."""
        wb = openpyxl.Workbook()
        wb.active.title = "Categorias"
        wb.active.append(["name"])

        # No products sheet — only categories
        prod_ws = find_sheet(wb, {"productos", "products"})
        assert prod_ws is None


class TestBuildHeaderMap:
    def test_standard_headers(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "sku", "category", "brand", "price", "cost", "stock"])

        header_map = build_header_map(ws)
        assert header_map["name"] == 0
        assert header_map["sku"] == 1
        assert header_map["category"] == 2
        assert header_map["brand"] == 3
        assert "selling_price" in header_map

    def test_spanish_headers(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["producto", "codigo", "categoria", "marca", "precio_venta", "costo", "cantidad"])

        header_map = build_header_map(ws)
        assert header_map["name"] == 0
        assert header_map["sku"] == 1
        assert header_map["category"] == 2
        assert header_map["brand"] == 3
